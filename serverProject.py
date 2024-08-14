import networkx as nx
import openpyxl
from flask import Flask, request, jsonify, render_template, redirect, url_for
import os
from openpyxl.reader.excel import load_workbook
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import PyPDF2
app = Flask(__name__)
arr_global=[]

@app.route('/fileExcel', methods=['POST'])
def file_excel():
    if 'file' not in request.files:
        return 'No file part'
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'
    file.save(file.filename)
    file_url = os.path.abspath(file.filename)
    workbook = load_workbook(filename=file_url)
    number_of_sheets = len(workbook.sheetnames)
    return {"file_url":file_url,"num_of_sheets":number_of_sheets}

@app.route('/report',  methods=['POST'])
def report():
    data = request.get_json()
    url = data.get('url')
    sheets = data.get('sheets')
    arr=[]
    for sheet in sheets:
        df = pd.read_excel(url, sheet_name=sheet['name'])
        column_sum = 0
        for column in sheet['columns']:
            if sheet['action'] == 'avg':
                column_sum += df[column].mean()
            if sheet['action'] == 'sum':
                column_sum += df[column].sum()
        arr.append(column_sum)
        global arr_global
        arr_global=arr
        print(arr_global)
    return jsonify(arr_global)
    # return redirect(url_for('convertPFD', values={"arr":arr}))

def convertPFD(values):
    print(values)
    c = canvas.Canvas("report.pdf", pagesize=letter)
    text = c.beginText(100, 700)
    text.setFont("Helvetica", 12)
    text.textLine("Integer Array Items:")
    for num in values:
        text.textLine(str(num))
    c.drawText(text)
    c.showPage()
    c.save()



def create_column_graph(data):
    arr_name_file=[]
    name_of_file=data[0]['name_of_file']
    name_of_sheet=data[1]['name_of_sheet']
    values=data[2]['values']
    sum_files = [100, 20]
    ind=0
    for file_name in name_of_file:
        plt.figure()
        plt.bar(name_of_sheet, values[ind])
        plt.xlabel('sum')
        plt.title(file_name)
        for i, value in enumerate(values[ind]):
            plt.text(i, value + 0.5, str(value), ha='center')
        file_name = f'{file_name}.png'
        plt.savefig(file_name)
        arr_name_file.append(file_name)
        ind+=1
    plt.figure()
    plt.bar(name_of_file, sum_files)
    plt.xlabel('average')
    plt.title("average of files")
    for i, value in enumerate(sum_files):
        plt.text(i, value + 0.5, str(value), ha='center')
    file_name = 'average.png'
    plt.savefig(file_name)
    arr_name_file.append(file_name)
    return arr_name_file

def convert_graphs_to_pdf(image_files):
    c = canvas.Canvas('graphs.pdf')
    for idx, image_file in enumerate(image_files):
        c.drawInlineImage(image_file, 100, 100, width=400, height=400)
        if idx != len(image_files) - 1:
            c.showPage()
    c.save()

graph_data1 = [
    {'name_of_file': ['file1', 'file2']},
    {'name_of_sheet': ['sheet1', 'sheet2', 'sheet3']},
    {'values': [[25, 40, 50], [25, 30, 50]]}
]

files=create_column_graph(graph_data1)
convert_graphs_to_pdf(files)

if __name__ == '__main__':
    app.run()
    convertPFD(arr_global)
